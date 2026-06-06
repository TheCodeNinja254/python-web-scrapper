"""
Hello,

Here is a quick guide before you edit this code.
Project Name: Manitoba Child Care Search — Full Scraper
==========================================
Scrapes all licensed child care facilities from childcaresearch.gov.mb.ca
and writes them to manitoba_childcare.xlsx.

Requirements:
    pip install playwright openpyxl
    python -m playwright install chromium
    pip install requests

    Run each of the above, one at a time, preferably

Usage:
    python mb_childcare_scraper.py                  # Full scrape → saves to manitoba_childcare.xlsx
    python mb_childcare_scraper.py --test           # Quick test: scrapes only 10 facilities
    python mb_childcare_scraper.py --diagnose   # Diagnose what API calls the site makes (run this first if broken)

Wishing you well,
Fredrick M
"""
from __future__ import annotations

import argparse
import json
import sys
import re
from datetime import datetime
from pathlib import Path

import requests
from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "manitoba_childcare.xlsx"
BASE_URL = "https://childcaresearch.gov.mb.ca"
SEARCH_URL = f"{BASE_URL}/en/Home/SearchChildCareFacilities"
LOCATIONS_URL = f"{BASE_URL}/en/Home/SearchedChildCareLocations"

BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

# ── Exact payload format the site uses (captured from browser network tab) ──
# All searchBy* fields set to true = search across all fields, no keyword filter
SEARCH_PAYLOAD = {
    "searchValue": "",
    "sortValue": "Facility Name - Ascending",
    "searchByCity": True,
    "searchByAddress": True,
    "searchByRegion": True,
    "searchByPostalCode": True,
    "SearchByArea": True,
    "SearchByNeighbourhood": True,
    "searchByFacilityName": True,
    "searchByLanguage": True,
    "searchByServicesDescription": True,
    "searchByEnrollmentPolicies": True,
    "showCentres": True,
    "showNurseries": True,
    "showHomes": True,
    "openWeekdays": False,
    "openWeekends": False,
    "openEvenings": False,
    "openOvernight": False,
    "infant": False,
    "preschool": False,
    "nursery": False,
    "schoolAge": False,
    "vacanciesOnly": False,
    "englishOnly": False,
    "frenchOnly": False,
    "both": False,
    "ccaEmployment": False,
    "eceIIEmployment": False,
    "eceIIIEmployment": False,
    "forProfit": False,
    "nonProfit": False,
    "funded": False,
    "nonFunded": False,
    "subsidized": False,
    "nonSubsidized": False,
    "locatedOnSchoolSite": False,
    "currentPage": 1,
    "pageSize": 10000,  # request all records at once
}


# ─────────────────────────────────────────────────────────────────────────────
# Session setup — loads the page in a real browser to get cookies + CSRF token
# ─────────────────────────────────────────────────────────────────────────────

def get_session(headed=False):
    """
    Visit the site in a headless browser to collect:
      - Session cookies
      - The RequestVerificationToken (ASP.NET antiforgery token)
      - The exact full request body the page sends (captured live)
    Returns (cookies_dict, csrf_token, full_captured_payload)
    """
    print("  Opening browser to establish session...")
    cookies_out = {}
    csrf_token = None
    captured_payload = None

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=not headed)
        ctx = browser.new_context(extra_http_headers={"User-Agent": BROWSER_UA})
        page = ctx.new_page()

        def on_request(request):
            nonlocal captured_payload
            if "SearchChildCareFacilities" in request.url and request.method == "POST":
                try:
                    body = json.loads(request.post_data or "{}")
                    # Keep the most complete payload we see
                    if len(body) > len(captured_payload or {}):
                        captured_payload = body
                except Exception:
                    pass

        page.on("request", on_request)

        try:
            page.goto(BASE_URL, wait_until="load", timeout=60000)
            print("  Waiting for page to fire API calls (20s)...")
            page.wait_for_timeout(15000)
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(5000)
        except Exception as e:
            print(f"  ⚠️  Navigation warning: {e}")

        # Collect cookies
        for c in ctx.cookies():
            cookies_out[c["name"]] = c["value"]

        # Find antiforgery token in rendered HTML
        html = page.content()
        for pattern in [
            r'__RequestVerificationToken["\s]+value=["\']([^"\']+)',
            r'"__RequestVerificationToken"\s*:\s*"([^"]+)"',
            r'name="__RequestVerificationToken"[^>]+value="([^"]+)"',
            r'value="([A-Za-z0-9+/=_\-]{50,})"',  # any long base64-ish value
        ]:
            m = re.search(pattern, html)
            if m:
                csrf_token = m.group(1)
                break

        # Also check cookies for the token (some ASP.NET apps put it there)
        for name, value in cookies_out.items():
            if "verif" in name.lower() or "csrf" in name.lower() or "xsrf" in name.lower():
                csrf_token = value
                break

        browser.close()

    print(f"  Cookies collected: {list(cookies_out.keys()) or 'none'}")
    print(f"  CSRF token: {'found ✅' if csrf_token else 'not found (will try without)'}")
    if captured_payload:
        print(f"  Live payload captured ✅ ({len(captured_payload)} fields)")

    return cookies_out, csrf_token, captured_payload


# ─────────────────────────────────────────────────────────────────────────────
# API call
# ─────────────────────────────────────────────────────────────────────────────

def call_api(cookies: dict, csrf_token: str | None, payload: dict) -> dict | list | None:
    """POST to SearchChildCareFacilities and return parsed JSON."""
    headers = {
        "User-Agent": BROWSER_UA,
        "Content-Type": "application/json",
        "Accept": "application/json, text/plain, */*",
        "Origin": BASE_URL,
        "Referer": BASE_URL + "/",
        "X-Requested-With": "XMLHttpRequest",
    }
    if csrf_token:
        # ASP.NET accepts the token under any of these header names
        headers["RequestVerificationToken"] = csrf_token
        headers["__RequestVerificationToken"] = csrf_token

    session = requests.Session()
    session.headers.update(headers)
    for name, value in cookies.items():
        session.cookies.set(name, value)

    try:
        resp = session.post(SEARCH_URL, json=payload, timeout=90)
        print(f"  HTTP {resp.status_code}")
        if resp.status_code == 200:
            return resp.json()
        else:
            print(f"  Error response: {resp.text[:300]}")
            return None
    except Exception as e:
        print(f"  ❌  Request failed: {e}")
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Parse API response → list of flat facility dicts
# ─────────────────────────────────────────────────────────────────────────────

def parse_facilities(data) -> list:
    """
    The API may return:
      - A list directly
      - {"facilities": [...]}
      - {"data": [...]}
      - {"results": [...]}
      - {"childCareFacilities": [...]}
    We try all common shapes.
    """
    if isinstance(data, list):
        records = data
    elif isinstance(data, dict):
        for key in ["facilities", "childCareFacilities", "data", "results", "items", "value"]:
            if key in data and isinstance(data[key], list):
                records = data[key]
                break
        else:
            # Fall back to the first list value found
            records = next((v for v in data.values() if isinstance(v, list)), [])
    else:
        return []

    return [flatten_facility(r) for r in records if isinstance(r, dict)]


def _get(d: dict, *keys, default="") -> str:
    """Try multiple key names, return first non-empty value."""
    for k in keys:
        v = d.get(k)
        if v is not None and str(v).strip() != "":
            return str(v).strip()
    return default


def flatten_facility(r: dict) -> dict:
    """Map one raw API record to our standard columns."""

    # Vacancies — may be nested under a 'vacancies' object or flat fields
    vac = r.get("vacancies") or r.get("vacancy") or {}
    if not isinstance(vac, dict):
        vac = {}

    # Employment — list of strings or dicts
    emp_raw = r.get("employmentOpportunities") or r.get("employment") or []
    if isinstance(emp_raw, list):
        emp = ", ".join(
            (e.get("name") or e.get("type") or str(e)) if isinstance(e, dict) else str(e)
            for e in emp_raw if e
        )
    else:
        emp = str(emp_raw).strip()

    # Operating times — may be list or flat bool fields
    times_list = r.get("operatingTimes") or []

    def is_open(label_fragments: list, bool_key: str = None) -> str:
        if bool_key and r.get(bool_key):
            return "Yes"
        for t in (times_list if isinstance(times_list, list) else []):
            name = (t.get("name") or t.get("type") or str(t)).lower() if isinstance(t, dict) else str(t).lower()
            if any(f in name for f in label_fragments):
                return "Yes"
        return "No"

    fid = _get(r, "childCareFacilityId", "facilityId", "id", "Id")
    profile_url = f"{BASE_URL}/en/Facility/Index?ChildCareFacilityId={fid}" if fid else ""

    return {
        "Facility ID": fid,
        "Facility Name": _get(r, "facilityName", "name", "title"),
        "Facility Type": _get(r, "facilityType", "type", "facilityTypeName"),
        "Business Type": _get(r, "businessType", "businessTypeName"),
        "Funding Type": _get(r, "fundingType", "fundingTypeName"),
        "Language": _get(r, "language", "languageOfCare", "languageName"),
        "Address": _get(r, "address", "fullAddress", "streetAddress"),
        "City": _get(r, "city", "cityTown", "town"),
        "Province": _get(r, "province", "provinceCode"),
        "Postal Code": _get(r, "postalCode", "postal"),
        "Region": _get(r, "region", "regionName"),
        "Area": _get(r, "area", "areaName"),
        "Neighbourhood": _get(r, "neighbourhood", "neighborhoodName", "neighborhoodName"),
        "Phone": _get(r, "phone", "phoneNumber", "telephone"),
        "Email": _get(r, "email", "emailAddress"),
        "Website": _get(r, "website", "websiteUrl", "url"),
        "Infant (0–2) Vacancies": _get(vac, "infant", "age0to2") or _get(r, "infantVacancies", "vacancyInfant",
                                                                         "age0to2Vacancies"),
        "Preschool (2–6) Vacancies": _get(vac, "preschool", "age2to6Preschool") or _get(r, "preschoolVacancies",
                                                                                        "vacancyPreschool"),
        "Nursery (2–6) Vacancies": _get(vac, "nursery", "age2to6Nursery") or _get(r, "nurseryVacancies",
                                                                                  "vacancyNursery"),
        "School Age (6–12) Vacancies": _get(vac, "schoolAge", "age6to12") or _get(r, "schoolAgeVacancies",
                                                                                  "vacancySchoolAge"),
        "Open Weekdays": is_open(["weekday"], "openWeekdays"),
        "Open Weekends": is_open(["weekend"], "openWeekends"),
        "Open Evenings": is_open(["evening"], "openEvenings"),
        "Open Overnight": is_open(["overnight"], "openOvernight"),
        "Employment Opportunities": emp,
        "Last Updated": _get(r, "lastUpdate", "lastUpdated", "updatedDate", "dateUpdated"),
        "Latitude": _get(r, "latitude", "lat"),
        "Longitude": _get(r, "longitude", "lng", "lon"),
        "Profile URL": profile_url,
        "Scraped At": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Diagnose mode
# ─────────────────────────────────────────────────────────────────────────────

def run_diagnose():
    print("\n🔬  DIAGNOSE MODE\n")
    cookies, csrf, live_payload = get_session()

    # Use the live-captured payload if we got one, else fall back to our known payload
    payload = live_payload or SEARCH_PAYLOAD
    print(f"\n📦  Using payload ({len(payload)} fields):")
    print(json.dumps(payload, indent=2))

    print(f"\n📥  Calling SearchChildCareFacilities...")
    data = call_api(cookies, csrf, payload)

    if data is None:
        print("\n❌  API returned an error.")
        return

    # Show structure
    print(f"\n✅  Response type: {type(data).__name__}")
    if isinstance(data, dict):
        print(f"   Top-level keys: {list(data.keys())}")
        for k, v in data.items():
            if isinstance(v, list):
                print(f"   '{k}': list of {len(v)} items")
                if v and isinstance(v[0], dict):
                    print(f"   First item keys: {list(v[0].keys())}")
                    print(f"\n   First record sample:\n{json.dumps(v[0], indent=4, default=str)[:2000]}")
    elif isinstance(data, list):
        print(f"   List of {len(data)} items")
        if data and isinstance(data[0], dict):
            print(f"   First item keys: {list(data[0].keys())}")
            print(f"\n   First record sample:\n{json.dumps(data[0], indent=4, default=str)[:2000]}")

    # Save full response
    out_file = "diagnose_response.json"
    Path(out_file).write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
    print(f"\n📄  Full response saved to: {out_file}")

    # Try parsing
    facilities = parse_facilities(data)
    print(f"\n✅  parse_facilities() found: {len(facilities)} facilities")
    if facilities:
        print(f"\n   Sample parsed row:\n{json.dumps(facilities[0], indent=4)[:1000]}")
    else:
        print("   ⚠️  0 facilities parsed — check diagnose_response.json and share it.")


# ─────────────────────────────────────────────────────────────────────────────
# Excel writer
# ─────────────────────────────────────────────────────────────────────────────

COLUMNS = [
    "Facility ID", "Facility Name", "Facility Type", "Business Type",
    "Funding Type", "Language",
    "Address", "City", "Province", "Postal Code",
    "Region", "Area", "Neighbourhood",
    "Phone", "Email", "Website",
    "Infant (0–2) Vacancies", "Preschool (2–6) Vacancies",
    "Nursery (2–6) Vacancies", "School Age (6–12) Vacancies",
    "Open Weekdays", "Open Weekends", "Open Evenings", "Open Overnight",
    "Employment Opportunities", "Last Updated",
    "Latitude", "Longitude", "Profile URL", "Scraped At",
]

COL_WIDTHS = {
    "Facility ID": 12, "Facility Name": 35, "Facility Type": 15,
    "Business Type": 15, "Funding Type": 15, "Language": 12,
    "Address": 38, "City": 18, "Province": 10, "Postal Code": 12,
    "Region": 18, "Area": 18, "Neighbourhood": 20,
    "Phone": 16, "Email": 28, "Website": 30,
    "Infant (0–2) Vacancies": 18, "Preschool (2–6) Vacancies": 22,
    "Nursery (2–6) Vacancies": 20, "School Age (6–12) Vacancies": 22,
    "Open Weekdays": 14, "Open Weekends": 14, "Open Evenings": 14, "Open Overnight": 14,
    "Employment Opportunities": 24, "Last Updated": 16,
    "Latitude": 12, "Longitude": 12, "Profile URL": 55, "Scraped At": 18,
}

HDR_FILL = PatternFill("solid", start_color="1F4E79")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
NORM_FILL = PatternFill("solid", start_color="FFFFFF")
BODY_FONT = Font(name="Arial", size=9)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_excel(facilities: list, output_path: str):
    if Path(output_path).exists():
        wb = load_workbook(output_path)
        ws = wb["Facilities"] if "Facilities" in wb.sheetnames else wb.active
        ws.delete_rows(2, ws.max_row)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Facilities"
        _add_instructions_sheet(wb)

    ws.row_dimensions[1].height = 30
    for ci, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 16)
    ws.freeze_panes = "A2"

    for ri, fac in enumerate(facilities, 2):
        fill = ALT_FILL if ri % 2 == 0 else NORM_FILL
        ws.row_dimensions[ri].height = 20
        for ci, col in enumerate(COLUMNS, 1):
            cell = ws.cell(row=ri, column=ci, value=fac.get(col, ""))
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = LEFT

    summary = len(facilities) + 2
    ws.cell(row=summary, column=1,
            value=f"Total facilities: {len(facilities)}").font = Font(name="Arial", bold=True, size=9)
    ws.cell(row=summary, column=len(COLUMNS),
            value=f"Last scraped: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(name="Arial", italic=True,
                                                                                            size=9)

    wb.save(output_path)
    print(f"\n✅  Saved {len(facilities)} facilities → {output_path}")


def _add_instructions_sheet(wb):
    info = wb.create_sheet("▶ How to Refresh")
    info.column_dimensions["A"].width = 85
    info["A1"] = "HOW TO REFRESH THIS SPREADSHEET"
    info["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    info["A1"].fill = PatternFill("solid", start_color="D6E4F0")
    steps = [
        "", "Run the scraper to refresh all data:", "",
        "    python mb_childcare_scraper.py",
        "", "Test with 10 facilities first:",
        "    python mb_childcare_scraper.py --test",
        "", "If something breaks, run the diagnose tool:",
        "    python mb_childcare_scraper.py --diagnose",
        "    This saves the raw API response to diagnose_response.json",
        "", "Source: https://childcaresearch.gov.mb.ca/",
        "Support: mbchildcaresearch@gov.mb.ca",
    ]
    for i, s in enumerate(steps, 2):
        info[f"A{i}"] = s
        info[f"A{i}"].font = Font(name="Arial", size=10)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Manitoba Child Care Scraper")
    parser.add_argument("--test", action="store_true", help="Only save first 10 facilities")
    parser.add_argument("--diagnose", action="store_true", help="Inspect API response structure")
    parser.add_argument("--headed", action="store_true", help="Show browser window while running")
    parser.add_argument("--output", default=OUTPUT_FILE, help="Output Excel file path")
    args = parser.parse_args()

    if args.diagnose:
        run_diagnose()
        return

    mode = "TEST (10 facilities)" if args.test else "FULL (all Manitoba)"
    print(f"\n🔍  Manitoba Child Care Scraper — {mode}")
    print(f"    Output: {args.output}\n")

    # ── Step 1: Get session ──────────────────────────────────────────────────
    print("Step 1/3 — Establishing browser session...")
    cookies, csrf, live_payload = get_session(headed=args.headed)

    # Use the live-captured payload (exact format the site sends) if available,
    # otherwise fall back to our hardcoded known-good payload
    payload = live_payload or SEARCH_PAYLOAD
    if live_payload:
        print("  Using live-captured payload ✅")
    else:
        print("  Using built-in payload (live capture not available)")

    # ── Step 2: Call the API ─────────────────────────────────────────────────
    print("\nStep 2/3 — Calling search API...")
    data = call_api(cookies, csrf, payload)

    if data is None:
        print("\n❌  API call failed.")
        print("    Run:  python mb_childcare_scraper.py --diagnose")
        sys.exit(1)

    # ── Step 3: Parse + write Excel ──────────────────────────────────────────
    print("\nStep 3/3 — Parsing results...")
    facilities = parse_facilities(data)

    if not facilities:
        print("❌  API returned data but 0 facilities could be parsed from it.")
        print("    Run:  python mb_childcare_scraper.py --diagnose")
        print("    This saves diagnose_response.json — share it and we can fix the parser.")
        sys.exit(1)

    if args.test:
        facilities = facilities[:10]
        print(f"    [TEST] Limited to {len(facilities)} facilities")

    print(f"    {len(facilities)} facilities parsed ✅")
    write_excel(facilities, args.output)


if __name__ == "__main__":
    main()